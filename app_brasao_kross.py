from io import BytesIO
from pathlib import Path
from copy import copy
import re
import traceback
import zipfile
import math

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


st.set_page_config(page_title="BRASÃO / KROSS → THOTH (PRO)", page_icon="📦", layout="wide")

BASE_DIR = Path(__file__).resolve().parent
IGNORE_NAMES = {"", "TOTAL", "TOTAIS", "SUBTOTAL", "SUB-TOTAL", "PRODUTO", "PRODUTOS"}

MODEL_CANDIDATES = {
    "BRASAO_FRUTAS": [
        "BRASAO - FRUTAS PRE PEDIDO BRANCO.xlsx",
        "BRASAO FRUTAS BRANCO.xlsx",
    ],
    "BRASAO_LEGUMES": [
        "BRASAO - LEGUMES PRE PEDIDO BRANCO.xlsx",
        "BRASAO LEGUMES BRANCO.xlsx",
    ],
    "KROSS_FRUTAS": [
        "KROSS - FRUTAS PRE PEDIDO BRANCO.xlsx",
        "KROSS - PRE PEDIDO FRUTAS BRANCO.xlsx",
    ],
    "KROSS_LEGUMES": [
        "KROSS - LEGUMES PRE PEDIDO BRANCO.xlsx",
    ],
}


# --- TRAVA DE SEGURANÇA ---
if base_file is None:
    st.info("👈 Por favor, carregue a sua BASE DE PRODUTOS (table 2.xlsx) no menu lateral para liberar o sistema.")
    st.stop() # Isso trava o aplicativo aqui até a base ser inserida

if not uploaded_files:
    st.info("👈 Agora selecione os arquivos de pedido dos clientes (Brasão, Kross, etc).")
    st.stop()

# --- LEITURA DA BASE ---
# A partir daqui, o sistema sabe que os arquivos foram carregados
# --- SUBSTITUI AS LINHAS 54 A 58 ---
# 1. Carrega a base usando o arquivo que você subiu no menu lateral
base_produtos = carregar_base(df_base)

# 2. Junta todos os arquivos de pedido que você subiu de uma vez só
lista_pedidos = []
for file in uploaded_files:
    df_temp = pd.read_excel(file)
    lista_pedidos.append(df_temp)

# Une tudo em uma tabela só para processar
pedidos_df = pd.concat(lista_pedidos, ignore_index=True)

# 3. Roda o seu motor de processamento original
resultados, estatisticas, sem_base_df = processar_pedidos(pedidos_df, base_produtos)
# --- PASSO 1: CARREGAR A BASE ---
st.sidebar.markdown("### 🟡 1. CARREGAR BASE PRODUTOS")
base_file = st.sidebar.file_uploader("Selecione sua base atualizada (table 2.xlsx)", type=["xlsx", "xls", "csv"], key="base")

st.sidebar.markdown("---") # Cria uma linha divisória para organizar

# --- PASSO 2: CARREGAR OS PEDIDOS ---
st.sidebar.markdown("### 🔵 2. SELECIONAR PEDIDOS")
uploaded_files = st.sidebar.file_uploader("Selecione os arquivos de Pedido (Excel)", type=["xlsx", "xls", "csv"], accept_multiple_files=True, key="pedidos")

STORE_RULES = [
    {
        "store_id": "BRASAO_FERNANDO",
        "group": "BRASAO",
        "col_key": "1",
        "display": "Brasão Fernando",
        "filename_signals": ["FERNANDO"],
        "signals": ["FERNANDO MACHADO", "CENTRO", "226"],
    },
    {
        "store_id": "BRASAO_JARDIM",
        "group": "BRASAO",
        "col_key": "2",
        "display": "Brasão Jardim",
        "filename_signals": ["JARDIM"],
        "signals": ["SAO PEDRO", "JARDIM AMERICA", "2199"],
    },
    {
        "store_id": "BRASAO_XAXIM",
        "group": "BRASAO",
        "col_key": "3",
        "display": "Brasão Xaxim",
        "filename_signals": ["XAXIM"],
        "signals": ["LUIZ LUNARDI", "XAXIM", "810"],
    },
    {
        "store_id": "BRASAO_AVENIDA",
        "group": "BRASAO",
        "col_key": "4",
        "display": "Brasão Avenida",
        "filename_signals": ["AVENIDA"],
        "signals": ["RIO DE JANEIRO", "CENTRO", "108", "CHAPECO"],
    },
    {
        "store_id": "BRASAO_CD",
        "group": "BRASAO_CD",
        "col_key": "CD",
        "display": "Brasão CD",
        "filename_signals": ["CD"],
        "signals": ["RUA GASPAR", "ELDORADO", "153"],
    },
    {
        "store_id": "KROSS_CHAPECO",
        "group": "KROSS",
        "col_key": "1",
        "display": "Kross Atacadista",
        "filename_signals": ["KROSS", "CHAPECO"],
        "signals": ["JOHN KENNEDY", "PASSO DOS FORTES", "550"],
    },
    {
        "store_id": "KROSS_XAXIM",
        "group": "KROSS",
        "col_key": "2",
        "display": "Kross Xaxim",
        "filename_signals": ["KROSS", "XAXIM"],
        "signals": ["AMELIO PANIZZI", "XAXIM"],
    },
]

ORDER_STOP_MARKERS = [
    "AGENDAR A ENTREGA",
    "PENDENCIAS DE MERCADORIAS",
    "TOTAL DO FORNECEDOR",
    "CONTATOS DO FORNECEDOR",
    "COMPRADOR",
    "TOTAIS",
    "VALOR TOTAL",
    "PESO TOTAL",
    "ORIG DEST TP CODIGO",
]

UNIT_PATTERNS = [
    (r"\bBDJ\b", "BDJ"),
    (r"\bBANDEJA\b", "BDJ"),
    (r"\bMACO\b", "UND"),
    (r"\bUNIDADE\b", "UND"),
    (r"\bUND\b", "UND"),
    (r"\bKG\b", "KG"),
]

REMOVAL_TOKENS = {
    "BRASAO", "KROSS", "FRUTA", "FRUTAS", "LEGUME", "LEGUMES",
    "DEMARCHI", "MARCHI", "SHELF", "DE", "DO", "DA", "DOS", "DAS",
    "KG", "KG.", "BDJ", "BANDEJA", "UND", "UNIDADE", "UN", "CX", "CAIXA"
}


def norm_text(v):
    if v is None:
        return ""
    t = str(v).strip()
    if t.lower() in {"nan", "none"}:
        return ""
    return " ".join(t.split())


def norm_key(v):
    txt = norm_text(v).upper()
    txt = txt.replace("Á", "A").replace("À", "A").replace("Ã", "A").replace("Â", "A")
    txt = txt.replace("É", "E").replace("Ê", "E")
    txt = txt.replace("Í", "I")
    txt = txt.replace("Ó", "O").replace("Õ", "O").replace("Ô", "O")
    txt = txt.replace("Ú", "U")
    txt = txt.replace("Ç", "C")
    return re.sub(r"\s+", " ", txt).strip()


def parse_number(v):
    if v is None:
        return None
    txt = norm_text(v)
    if not txt:
        return None
    txt = txt.replace(".", "").replace(",", ".")
    txt = re.sub(r"[^0-9.-]", "", txt)
    if txt in {"", ".", "-", "-."}:
        return None
    try:
        val = float(txt)
        if math.isnan(val):
            return None
        return val
    except Exception:
        return None


def safe_positive(v):
    num = parse_number(v)
    return num if num and num > 0 else None


def ceil_div(qtd, base):
    qtd = parse_number(qtd)
    base = parse_number(base)
    if qtd is None or base is None or base <= 0:
        return 0
    return int(math.ceil(qtd / base))


def resolve_existing_file(candidates):
    for name in candidates:
        path = BASE_DIR / name
        if path.exists():
            return path
    return None


def resolve_fixed_paths():
    models = {}
    missing = []
    for key, candidates in MODEL_CANDIDATES.items():
        path = resolve_existing_file(candidates)
        if path is None:
            missing.append(candidates[0])
        else:
            models[key] = path

    base_file = resolve_existing_file(BASE_FILE_CANDIDATES)
    if base_file is None:
        missing.append(BASE_FILE_CANDIDATES[0])

    if missing:
        raise FileNotFoundError("Arquivos fixos não encontrados no repositório: " + ", ".join(missing))

    return base_file, models


@st.cache_data(show_spinner=False)
def load_base_from_disk(path_str: str) -> pd.DataFrame:
    file_path = Path(path_str)
    if not file_path.exists():
        raise FileNotFoundError(f"Base de produtos não encontrada: {file_path.name}")

    df = pd.read_excel(file_path, sheet_name="BASE_PRODUTOS")
    cols_map = {c: norm_key(c) for c in df.columns}

    def pick(col_options, required=False):
        for original, key in cols_map.items():
            if key in col_options:
                return original
        if required:
            raise ValueError(f"Coluna obrigatória não encontrada na base: {col_options}")
        return None

    col_categoria = pick({"CATEGORIA"}, required=True)
    col_produto = pick({"PRODUTO_BASE", "PRODUTO", "ITEM", "DESCRICAO", "DESCRICAO BASE"}, required=True)
    col_sinonimos = pick({"SINONIMOS", "SINONIMO", "APELIDOS"})
    col_codigo = pick({"CODIGO", "COD", "COD ITEM"})
    col_cod_forn = pick({"COD FORN", "COD_FORN", "CODFORN", "CODIGO FORNECEDOR"})
    col_modo = pick({"MODO_CONVERSAO", "MODO CONVERSAO", "MODO"})
    col_peso = pick({"PESO_CAIXA", "PESO CAIXA", "KG POR CAIXA"})
    col_itens = pick({"ITENS_POR_CAIXA", "ITENS POR CAIXA", "UN POR CAIXA", "UND POR CAIXA"})
    col_bdj = pick({"BANDEJAS_POR_CAIXA", "BANDEJAS POR CAIXA", "BDJ POR CAIXA"})
    col_status = pick({"STATUS_BASE", "STATUS"})
    col_valid = pick({"VALIDACAO", "VALIDAÇÃO"})

    rows = []
    for _, r in df.iterrows():
        produto = norm_text(r.get(col_produto))
        if not produto:
            continue

        sinonimos_raw = norm_text(r.get(col_sinonimos)) if col_sinonimos else ""
        sinonimos = []
        if sinonimos_raw:
            for part in re.split(r"[|;]", sinonimos_raw):
                p = norm_text(part)
                if p:
                    sinonimos.append(p)

        produto_key = norm_key(produto)
        produto_tokens = [t for t in re.split(r"\W+", produto_key) if t and t not in REMOVAL_TOKENS and len(t) > 1]
        sinonimos_key = [norm_key(x) for x in sinonimos]

        rows.append({
            "categoria": norm_key(r.get(col_categoria)),
            "produto_base": produto,
            "produto_key": produto_key,
            "produto_tokens": produto_tokens,
            "sinonimos": sinonimos,
            "sinonimos_key": sinonimos_key,
            "codigo": norm_text(r.get(col_codigo)) if col_codigo else "",
            "cod_forn": norm_text(r.get(col_cod_forn)) if col_cod_forn else "",
            "modo": norm_key(r.get(col_modo)) if col_modo else "",
            "peso_caixa": safe_positive(r.get(col_peso)) if col_peso else None,
            "itens_por_caixa": safe_positive(r.get(col_itens)) if col_itens else None,
            "bandejas_por_caixa": safe_positive(r.get(col_bdj)) if col_bdj else None,
            "status_base": norm_text(r.get(col_status)) if col_status else "",
            "validacao": norm_text(r.get(col_valid)) if col_valid else "",
        })

    base_df = pd.DataFrame(rows)
    if base_df.empty:
        raise ValueError("A base de produtos está vazia.")
    return base_df


@st.cache_data(show_spinner=False)
def pdf_to_text(file_bytes: bytes) -> str:
    if pdfplumber is None:
        raise RuntimeError("Falta instalar pdfplumber no ambiente. Adicione em requirements.txt")

    all_text = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            all_text.append(txt)
    return "\n".join(all_text)


def identify_store(pdf_text: str, file_name: str = ""):
    text = norm_key(pdf_text)
    name = norm_key(Path(file_name).stem)

    if "CD" in name and "KROSS" not in name:
        return next(r for r in STORE_RULES if r["store_id"] == "BRASAO_CD")

    if "KROSS" in name and "XAXIM" in name:
        return next(r for r in STORE_RULES if r["store_id"] == "KROSS_XAXIM")
    if "KROSS" in name and ("CHAPECO" in name or "ATACADO" in name or "C.O" in name or "CO" in name):
        return next(r for r in STORE_RULES if r["store_id"] == "KROSS_CHAPECO")
    if "FERNANDO" in name:
        return next(r for r in STORE_RULES if r["store_id"] == "BRASAO_FERNANDO")
    if "JARDIM" in name:
        return next(r for r in STORE_RULES if r["store_id"] == "BRASAO_JARDIM")
    if "AVENIDA" in name:
        return next(r for r in STORE_RULES if r["store_id"] == "BRASAO_AVENIDA")
    if "XAXIM" in name and "KROSS" not in name:
        return next(r for r in STORE_RULES if r["store_id"] == "BRASAO_XAXIM")

    for rule in STORE_RULES:
        if all(signal in text for signal in rule["signals"]):
            return rule

    raise ValueError(f"Não consegui identificar a loja/unidade no PDF: {file_name}")
def detect_unit(desc: str) -> str:
    key = norm_key(desc)
    for pattern, unit in UNIT_PATTERNS:
        if re.search(pattern, key):
            return unit
    return "CX"


def clean_desc(desc: str) -> str:
    return re.sub(r"\s+", " ", norm_text(desc)).strip()


def normalize_desc_for_match(desc: str) -> str:
    key = norm_key(desc)
    key = re.sub(r"\b\d+(?:[.,]\d+)?\s*G\b", " ", key)
    key = re.sub(r"\b\d+(?:[.,]\d+)?\s*KG\b", " ", key)
    key = re.sub(r"\bSHELF\s*\d+\b", " ", key)
    key = re.sub(r"\b(IMPORTADO|IMPORTADA|NACIONAL|GRECIA|GRECIA)\b", lambda m: m.group(1), key)

    words = []
    for token in re.split(r"\W+", key):
        if not token:
            continue
        if token in REMOVAL_TOKENS:
            continue
        if re.fullmatch(r"\d+", token):
            continue
        words.append(token)
    return " ".join(words).strip()
def is_stop_line(line: str) -> bool:
    key = norm_key(line)
    return any(marker in key for marker in ORDER_STOP_MARKERS)


def parse_order_items(pdf_text: str) -> pd.DataFrame:
    lines = [norm_text(x) for x in pdf_text.splitlines() if norm_text(x)]
    item_lines = []
    start_collecting = False

    row_pattern = re.compile(
        r"^(?P<codigo>\d[\d.,]*)\s+"
        r"(?P<cod_forn>[\d,./-]+)\s+"
        r"(?P<descricao>.+?)\s+"
        r"(?P<quant>\d+[\d.,]*)\s+"
        r"(?P<qtde_emb>\d+[\d.,]*)\s+"
        r"(?P<pr_unit>\d+[\d.,]*)\s+"
        r"(?P<vl_total>\d+[\d.,]*)$"
    )

    for line in lines:
        key = norm_key(line)
        if "CODIGO COD FORN DESCRICAO" in key or "CODIGO COD FORN DESCRI" in key:
            start_collecting = True
            continue

        if not start_collecting:
            continue

        if is_stop_line(line):
            break

        match = row_pattern.match(line)
        if not match:
            continue

        descricao = clean_desc(match.group("descricao"))
        item_lines.append({
            "CodigoPedido": norm_text(match.group("codigo")),
            "CodFornPedido": norm_text(match.group("cod_forn")),
            "DescricaoOriginal": descricao,
            "DescricaoNormalizada": normalize_desc_for_match(descricao),
            "Qtde": parse_number(match.group("quant")) or 0,
            "QtdeEmb": parse_number(match.group("qtde_emb")) or 0,
            "PrecoPedido": parse_number(match.group("pr_unit")),
            "ValorTotal": parse_number(match.group("vl_total")),
            "UnidadeDetectada": detect_unit(descricao),
        })

    df = pd.DataFrame(item_lines)
    if df.empty:
        raise ValueError("Nenhum item válido foi extraído do PDF. Verifique se o layout do pedido segue o padrão atual.")
    return df


def match_base_item(desc: str, base_df: pd.DataFrame):
    key = normalize_desc_for_match(desc)
    if not key:
        return None

    exact = base_df[base_df["produto_key"] == key]
    if not exact.empty:
        return exact.iloc[0]

    # 1) sinônimo exato
    for _, row in base_df.iterrows():
        if key in row["sinonimos_key"] or row["produto_key"] == key:
            return row

    # 2) sinônimo por contenção
    candidates = []
    key_tokens = [t for t in re.split(r"\W+", key) if t and t not in REMOVAL_TOKENS and len(t) > 1]
    key_set = set(key_tokens)

    for _, row in base_df.iterrows():
        prod_key = row["produto_key"]
        syns = row["sinonimos_key"]

        score = 0
        if prod_key and (prod_key in key or key in prod_key):
            score += 100

        for syn in syns:
            if syn and (syn in key or key in syn):
                score += 90
                break

        row_tokens = set(row["produto_tokens"])
        overlap = len(key_set.intersection(row_tokens))
        if overlap:
            score += overlap * 10

        if score > 0:
            candidates.append((score, len(prod_key), row))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best = candidates[0][2]

    # exige pelo menos 2 tokens em comum para nomes grandes, ou score alto de contenção
    top_score = candidates[0][0]
    if top_score >= 90:
        return best

    overlap = len(key_set.intersection(set(best["produto_tokens"])))
    if len(key_tokens) <= 2 and overlap >= 1:
        return best
    if len(key_tokens) >= 3 and overlap >= 2:
        return best

    return None
def choose_quantity(qtd: float, qtd_emb: float, modo: str, detected_unit: str) -> float:
    qtd = parse_number(qtd) or 0
    qtd_emb = parse_number(qtd_emb) or 0
    modo = norm_key(modo)
    detected_unit = norm_key(detected_unit)

    if modo == "CAIXA":
        return qtd if qtd > 0 else qtd_emb

    # para KG/UND/BDJ, quando o pdf vier no formato 1 | 400, usar o maior valor
    if qtd_emb > qtd:
        return qtd_emb
    if qtd > 0:
        return qtd
    return qtd_emb
def convert_to_boxes(row, base_row):
    modo = norm_key(base_row.get("modo", ""))
    unit = norm_key(row.get("UnidadeDetectada", ""))
    qtd = row.get("Qtde")
    qtd_emb = row.get("QtdeEmb")
    qtd_real = choose_quantity(qtd, qtd_emb, modo, unit)

    if modo == "CAIXA":
        return (int(math.ceil(qtd_real)) if qtd_real > 0 else 0), qtd_real, None

    if modo == "PESO":
        peso = safe_positive(base_row.get("peso_caixa"))
        if not peso:
            return 0, qtd_real, "BASE INCOMPLETA: peso_caixa"
        return ceil_div(qtd_real, peso), qtd_real, None

    if modo in {"UNIDADE", "UND"}:
        itens = safe_positive(base_row.get("itens_por_caixa"))
        if not itens:
            return 0, qtd_real, "BASE INCOMPLETA: itens_por_caixa"
        return ceil_div(qtd_real, itens), qtd_real, None

    if modo in {"BANDEJA", "BDJ"}:
        bdj = safe_positive(base_row.get("bandejas_por_caixa"))
        if not bdj:
            return 0, qtd_real, "BASE INCOMPLETA: bandejas_por_caixa"
        return ceil_div(qtd_real, bdj), qtd_real, None

    # fallback pela unidade detectada, mas também bloqueando base incompleta
    if unit == "KG":
        peso = safe_positive(base_row.get("peso_caixa"))
        if not peso:
            return 0, qtd_real, "BASE INCOMPLETA: peso_caixa"
        return ceil_div(qtd_real, peso), qtd_real, None
    if unit == "UND":
        itens = safe_positive(base_row.get("itens_por_caixa"))
        if not itens:
            return 0, qtd_real, "BASE INCOMPLETA: itens_por_caixa"
        return ceil_div(qtd_real, itens), qtd_real, None
    if unit == "BDJ":
        bdj = safe_positive(base_row.get("bandejas_por_caixa"))
        if not bdj:
            return 0, qtd_real, "BASE INCOMPLETA: bandejas_por_caixa"
        return ceil_div(qtd_real, bdj), qtd_real, None
    if unit == "CX":
        return (int(math.ceil(qtd_real)) if qtd_real > 0 else 0), qtd_real, None

    return 0, qtd_real, "BASE INCOMPLETA: modo/unidade sem regra"
def transform_items(order_df: pd.DataFrame, store_rule: dict, base_df: pd.DataFrame):
    out_rows = []
    errors = []

    for _, row in order_df.iterrows():
        base_row = match_base_item(row["DescricaoOriginal"], base_df)
        if base_row is None:
            errors.append({
                "loja": store_rule["display"],
                "produto": row["DescricaoOriginal"],
                "erro": "Produto não encontrado na base",
            })
            continue

        caixas, qtd_real, motivo = convert_to_boxes(row, base_row)
        if caixas <= 0:
            detalhe = motivo or f"Falha na conversão para caixa ({row['UnidadeDetectada']})"
            errors.append({
                "loja": store_rule["display"],
                "produto": row["DescricaoOriginal"],
                "erro": f"{detalhe} [qtd={row['Qtde']} emb={row['QtdeEmb']} qtd_real={qtd_real}]",
            })
            continue

        out_rows.append({
            "Grupo": store_rule["group"],
            "LojaKey": store_rule["col_key"],
            "LojaNome": store_rule["display"],
            "Categoria": norm_key(base_row["categoria"]),
            "ProdutoModelo": base_row["produto_base"],
            "ProdutoKey": base_row["produto_key"],
            "Caixas": caixas,
            "CodigoPedido": row["CodigoPedido"],
            "CodFornPedido": row["CodFornPedido"],
            "PrecoPedido": row["PrecoPedido"],
        })

    out_df = pd.DataFrame(out_rows)
    errors_df = pd.DataFrame(errors)
    return out_df, errors_df
def group_to_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    p = df.pivot_table(index="ProdutoModelo", columns="LojaKey", values="Caixas", aggfunc="sum", fill_value=0)
    p.columns = [str(c) for c in p.columns]
    return p


def product_rows(ws):
    rows = {}
    for row in range(3, ws.max_row + 1):
        prod = norm_text(ws.cell(row, 1).value)
        if prod and norm_key(prod) not in IGNORE_NAMES:
            rows[norm_key(prod)] = row
    return rows


def copy_row_style(ws, src_row, dst_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def extract_store_number_from_model(v1, v2):
    c1 = norm_key(v1)
    c2 = norm_key(v2)
    combo = f"{c1} {c2}"
    if "LOJA 1" in combo or re.fullmatch(r"1", c2):
        return "1"
    if "LOJA 2" in combo or re.fullmatch(r"2", c2):
        return "2"
    if "LOJA 3" in combo or re.fullmatch(r"3", c2):
        return "3"
    if "LOJA 4" in combo or re.fullmatch(r"4", c2):
        return "4"
    return None


def model_map(ws):
    store_to_col = {}
    total_col = None
    for col in range(2, ws.max_column + 1):
        line1 = ws.cell(1, col).value
        line2 = ws.cell(2, col).value
        top = norm_key(line1)
        second = norm_key(line2)

        if "TOTAL" in top or "TOTAL" in second:
            total_col = col
            continue

        key = extract_store_number_from_model(line1, line2)
        if key:
            store_to_col[key] = col

    return store_to_col, total_col


def model_map_kross(ws):
    store_to_col = {}
    total_col = None
    for col in range(2, ws.max_column + 1):
        line1 = norm_key(ws.cell(1, col).value)
        line2 = norm_key(ws.cell(2, col).value)
        combo = f"{line1} {line2}"

        if "TOTAL" in combo:
            total_col = col
            continue
        if "ATACADISTA" in combo or re.fullmatch(r"1", line2):
            store_to_col["1"] = col
        elif "XAXIM" in combo or re.fullmatch(r"2", line2):
            store_to_col["2"] = col

    return store_to_col, total_col


def write_output(model_path: Path, data: pd.DataFrame, model_type: str) -> bytes:
    wb = load_workbook(str(model_path))
    ws = wb.active

    if model_type == "KROSS":
        stores, total_col = model_map_kross(ws)
    else:
        stores, total_col = model_map(ws)

    prod_map = product_rows(ws)
    cols_to_clear = list(stores.values())
    if total_col:
        cols_to_clear.append(total_col)

    for row in range(3, ws.max_row + 1):
        if norm_text(ws.cell(row, 1).value):
            for col in cols_to_clear:
                ws.cell(row, col).value = None

    used = set()

    for prod in data.index.tolist():
        key = norm_key(prod)
        if key not in prod_map:
            continue

        row = prod_map[key]
        used.add(key)
        row_total = 0

        for loja in data.columns:
            if loja in stores:
                val = float(data.loc[prod, loja])
                if val:
                    ws.cell(row, stores[loja]).value = val
                    row_total += val

        if total_col:
            ws.cell(row, total_col).value = row_total if row_total else None

    missing = [prod for prod in data.index.tolist() if norm_key(prod) not in used]
    if missing:
        last_filled = max(prod_map.values()) if prod_map else 3
        style_row = last_filled
        current_row = last_filled + 1

        for prod in missing:
            copy_row_style(ws, style_row, current_row)
            ws.cell(current_row, 1).value = prod

            row_total = 0
            for loja in data.columns:
                if loja in stores:
                    val = float(data.loc[prod, loja])
                    if val:
                        ws.cell(current_row, stores[loja]).value = val
                        row_total += val

            if total_col:
                ws.cell(current_row, total_col).value = row_total if row_total else None
            current_row += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def build_prices_sheet(df: pd.DataFrame) -> bytes:
    if df.empty:
        out = BytesIO()
        pd.DataFrame(columns=["CODIGO", "COD_FORN", "PRODUTO", "PRECO"]).to_excel(out, index=False)
        out.seek(0)
        return out.getvalue()

    base = df[["ProdutoModelo", "CodigoPedido", "CodFornPedido", "PrecoPedido"]].drop_duplicates(subset=["ProdutoModelo"])
    base = base.sort_values("ProdutoModelo")

    rows = []
    for _, r in base.iterrows():
        codigo = norm_text(r["CodigoPedido"])
        cod_num = "".join(filter(str.isdigit, codigo)) if codigo else ""
        rows.append({
            "CODIGO": f"'{cod_num}" if cod_num else "",
            "COD_FORN": norm_text(r["CodFornPedido"]),
            "PRODUTO": norm_text(r["ProdutoModelo"]),
            "PRECO": r["PrecoPedido"],
        })

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="PRECOS", index=False)
        ws = writer.sheets["PRECOS"]
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 12
    out.seek(0)
    return out.getvalue()


def build_cd_workbook(frutas_matrix: pd.DataFrame, legumes_matrix: pd.DataFrame) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        (frutas_matrix if not frutas_matrix.empty else pd.DataFrame()).to_excel(writer, sheet_name="FRUTAS")
        (legumes_matrix if not legumes_matrix.empty else pd.DataFrame()).to_excel(writer, sheet_name="LEGUMES")
    out.seek(0)
    return out.getvalue()


def build_zip(files_dict: dict) -> bytes:
    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, content in files_dict.items():
            zf.writestr(name, content)
    out.seek(0)
    return out.getvalue()


st.title("📦 BRASÃO / KROSS → THOTH (PRO)")
st.caption("Suba apenas os PDFs. A base e os modelos já ficam fixos dentro do sistema.")

with st.sidebar:
    st.subheader("Como usar")
    st.write("1. Envie todos os PDFs do dia.")
    st.write("2. Clique em PROCESSAR.")
    st.write("3. Baixe o ZIP com os arquivos finais.")
    st.info("Fluxos suportados: Brasão lojas, Brasão CD e Kross.")
    st.info("Brasão Fernando = Loja 1 | Jardim = Loja 2 | Xaxim = Loja 3 | Avenida = Loja 4.")
    st.info("Kross Atacadista = Loja 1 | Kross Xaxim = Loja 2.")
    st.info(f"Base preferida: {BASE_FILE_CANDIDATES[0]}")

pdf_files = st.file_uploader("Pedidos PDF", type=["pdf"], accept_multiple_files=True)

if st.button("PROCESSAR", use_container_width=True, type="primary"):
    if not pdf_files:
        st.error("Envie ao menos um PDF.")
    else:
        try:
            base_file, models = resolve_fixed_paths()
            base_df = load_base_from_disk(str(base_file))

            transformed_parts = []
            all_errors = []
            identified = []
            seen_store_ids = set()

            for pdf in pdf_files:
                text = pdf_to_text(pdf.getvalue())
                store_rule = identify_store(text, pdf.name)

                if store_rule["store_id"] in seen_store_ids:
                    all_errors.append(pd.DataFrame([{
                        "loja": store_rule["display"],
                        "produto": pdf.name,
                        "erro": "PDF duplicado da mesma unidade ignorado",
                    }]))
                    continue

                seen_store_ids.add(store_rule["store_id"])
                order_df = parse_order_items(text)
                transformed_df, errors_df = transform_items(order_df, store_rule, base_df)

                if not transformed_df.empty:
                    transformed_parts.append(transformed_df)
                if not errors_df.empty:
                    all_errors.append(errors_df)

                identified.append({
                    "arquivo": pdf.name,
                    "loja": store_rule["display"],
                    "grupo": store_rule["group"],
                    "itens_extraidos": len(order_df),
                    "itens_convertidos": len(transformed_df),
                })

            if not transformed_parts:
                raise ValueError("Nenhum item foi convertido. Verifique a base e os PDFs.")

            all_data = pd.concat(transformed_parts, ignore_index=True)
            errors_data = pd.concat(all_errors, ignore_index=True) if all_errors else pd.DataFrame(columns=["loja", "produto", "erro"])
            identified_df = pd.DataFrame(identified)

            brasao_df = all_data[all_data["Grupo"] == "BRASAO"].copy()
            kross_df = all_data[all_data["Grupo"] == "KROSS"].copy()
            cd_df = all_data[all_data["Grupo"] == "BRASAO_CD"].copy()

            brasao_frutas = brasao_df[brasao_df["Categoria"] == "FRUTAS"]
            brasao_legumes = brasao_df[brasao_df["Categoria"] == "LEGUMES"]
            kross_frutas = kross_df[kross_df["Categoria"] == "FRUTAS"]
            kross_legumes = kross_df[kross_df["Categoria"] == "LEGUMES"]
            cd_frutas = cd_df[cd_df["Categoria"] == "FRUTAS"]
            cd_legumes = cd_df[cd_df["Categoria"] == "LEGUMES"]

            brasao_frutas_matrix = group_to_matrix(brasao_frutas)
            brasao_legumes_matrix = group_to_matrix(brasao_legumes)
            kross_frutas_matrix = group_to_matrix(kross_frutas)
            kross_legumes_matrix = group_to_matrix(kross_legumes)
            cd_frutas_matrix = group_to_matrix(cd_frutas)
            cd_legumes_matrix = group_to_matrix(cd_legumes)

            files_to_zip = {
                "BRASAO_FRUTAS_Thoth.xlsx": write_output(models["BRASAO_FRUTAS"], brasao_frutas_matrix, "BRASAO"),
                "BRASAO_LEGUMES_Thoth.xlsx": write_output(models["BRASAO_LEGUMES"], brasao_legumes_matrix, "BRASAO"),
                "KROSS_FRUTAS_Thoth.xlsx": write_output(models["KROSS_FRUTAS"], kross_frutas_matrix, "KROSS"),
                "KROSS_LEGUMES_Thoth.xlsx": write_output(models["KROSS_LEGUMES"], kross_legumes_matrix, "KROSS"),
                "BRASAO_CD.xlsx": build_cd_workbook(cd_frutas_matrix, cd_legumes_matrix),
                "BRASAO_PRECOS.xlsx": build_prices_sheet(brasao_df),
                "KROSS_PRECOS.xlsx": build_prices_sheet(kross_df),
                "BRASAO_CD_PRECOS.xlsx": build_prices_sheet(cd_df),
            }

            missing_units = []
            expected_ids = {
                "BRASAO_FERNANDO", "BRASAO_JARDIM", "BRASAO_XAXIM", "BRASAO_AVENIDA",
                "BRASAO_CD", "KROSS_CHAPECO", "KROSS_XAXIM"
            }
            for rule in STORE_RULES:
                if rule["store_id"] in expected_ids and rule["store_id"] not in seen_store_ids:
                    missing_units.append({"loja": rule["display"], "produto": "", "erro": "PDF da unidade não enviado"})

            if missing_units:
                missing_df = pd.DataFrame(missing_units)
                errors_data = pd.concat([errors_data, missing_df], ignore_index=True)

            err_out = BytesIO()
            with pd.ExcelWriter(err_out, engine="openpyxl") as writer:
                identified_df.to_excel(writer, sheet_name="ARQUIVOS", index=False)
                errors_data.to_excel(writer, sheet_name="ERROS", index=False)
            err_out.seek(0)
            files_to_zip["LOG_PROCESSAMENTO.xlsx"] = err_out.getvalue()

            zip_bytes = build_zip(files_to_zip)

            st.success("Processamento concluído com sucesso.")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("PDFs enviados", len(pdf_files))
            c2.metric("Unidades identificadas", len(identified_df))
            c3.metric("Itens convertidos", len(all_data))
            c4.metric("Ocorrências no log", len(errors_data))

            st.subheader("Arquivos processados")
            st.dataframe(identified_df, use_container_width=True)

            if not errors_data.empty:
                st.subheader("Log de ocorrências")
                st.dataframe(errors_data, use_container_width=True)

            st.download_button(
                "Baixar ZIP final",
                zip_bytes,
                file_name="THOTH_BRASAO_KROSS_PRO.zip",
                mime="application/zip",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"Erro ao processar: {e}")
            with st.expander("Ver detalhes do erro"):
                st.code(traceback.format_exc())
